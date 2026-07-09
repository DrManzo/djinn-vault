"""
accounting.py — Financial reporting layer for Djinn Shop.
Spec by Marcus (Perplexity). Implementation by Claude.

Reads from operational tables: orders, order_items, customers, ledger.
Writes to accounting snapshot tables: invoices, income_statements,
balance_sheets, monthly_reports.

All monetary values: USD float.
All dates: ISO 8601 strings (YYYY-MM-DD or YYYY-MM-DDTHH:MM:SSZ).
"""

import sqlite3
import json
import csv
import uuid
import datetime
import pathlib
from typing import Optional

import sys
_SHOP = pathlib.Path(__file__).parent
_PRINTER = _SHOP.parent
if str(_PRINTER) not in sys.path:
    sys.path.insert(0, str(_PRINTER))

from shop.db import get_db, SHOP_DIR, decrypt

EXPORTS_DIR = SHOP_DIR / "exports"

# ── Printer depreciation config ───────────────────────────────────────────────
PRINTER_PURCHASE_PRICE = 399.0
PRINTER_PURCHASE_DATE  = "2026-01-01"    # update at setup
PRINTER_USEFUL_LIFE_YR = 3
MACHINE_RATE_PER_HOUR  = 0.20
DEFAULT_LABOR_COST     = 6.67            # per job flat rate

# ── Schema additions ──────────────────────────────────────────────────────────
ACCOUNTING_SCHEMA = """
-- Extend customers table with acquisition channel + status
-- (ALTER TABLE is safe to re-run — fails silently if column exists)

ALTER TABLE customers ADD COLUMN channel TEXT DEFAULT 'discord';
ALTER TABLE customers ADD COLUMN customer_status TEXT DEFAULT 'active';

CREATE TABLE IF NOT EXISTS invoices (
    invoice_id       TEXT    PRIMARY KEY,
    order_id         TEXT    NOT NULL,
    customer_id      INTEGER NOT NULL,
    issued_date      TEXT    NOT NULL,
    due_date         TEXT,
    total_charged    REAL    NOT NULL DEFAULT 0.0,
    amount_paid      REAL    DEFAULT 0.0,
    payment_method   TEXT,
    paid_date        TEXT,
    status           TEXT    NOT NULL DEFAULT 'unpaid',
    FOREIGN KEY (order_id)    REFERENCES orders(id),
    FOREIGN KEY (customer_id) REFERENCES customers(id)
);

CREATE TABLE IF NOT EXISTS income_statements (
    id               INTEGER PRIMARY KEY AUTOINCREMENT,
    period_start     TEXT    NOT NULL,
    period_end       TEXT    NOT NULL,
    gross_revenue    REAL    DEFAULT 0.0,
    material_cogs    REAL    DEFAULT 0.0,
    machine_cogs     REAL    DEFAULT 0.0,
    labor_cogs       REAL    DEFAULT 0.0,
    design_cogs      REAL    DEFAULT 0.0,
    total_cogs       REAL    DEFAULT 0.0,
    gross_profit     REAL    DEFAULT 0.0,
    gross_margin_pct REAL    DEFAULT 0.0,
    platform_fees    REAL    DEFAULT 0.0,
    supply_expense   REAL    DEFAULT 0.0,
    operating_expense REAL   DEFAULT 0.0,
    net_income       REAL    DEFAULT 0.0,
    generated_at     TEXT    NOT NULL
);

CREATE TABLE IF NOT EXISTS balance_sheets (
    id               INTEGER PRIMARY KEY AUTOINCREMENT,
    snapshot_date    TEXT    NOT NULL,
    cash             REAL    DEFAULT 0.0,
    accounts_receivable REAL DEFAULT 0.0,
    inventory_value  REAL    DEFAULT 0.0,
    equipment_value  REAL    DEFAULT 0.0,
    total_assets     REAL    DEFAULT 0.0,
    accounts_payable REAL    DEFAULT 0.0,
    total_liabilities REAL   DEFAULT 0.0,
    owners_equity    REAL    DEFAULT 0.0,
    generated_at     TEXT    NOT NULL
);

CREATE TABLE IF NOT EXISTS monthly_reports (
    report_month       TEXT  PRIMARY KEY,
    jobs_completed     INTEGER DEFAULT 0,
    jobs_failed        INTEGER DEFAULT 0,
    failure_rate_pct   REAL    DEFAULT 0.0,
    gross_revenue      REAL    DEFAULT 0.0,
    total_cogs         REAL    DEFAULT 0.0,
    gross_profit       REAL    DEFAULT 0.0,
    gross_margin_pct   REAL    DEFAULT 0.0,
    net_income         REAL    DEFAULT 0.0,
    new_customers      INTEGER DEFAULT 0,
    repeat_customers   INTEGER DEFAULT 0,
    avg_job_value      REAL    DEFAULT 0.0,
    filament_used_grams REAL   DEFAULT 0.0,
    machine_hours      REAL    DEFAULT 0.0,
    top_customer_id    INTEGER,
    generated_at       TEXT    NOT NULL
);

CREATE INDEX IF NOT EXISTS idx_invoices_order    ON invoices(order_id);
CREATE INDEX IF NOT EXISTS idx_invoices_customer ON invoices(customer_id);
CREATE INDEX IF NOT EXISTS idx_invoices_status   ON invoices(status);
"""


def init_accounting(conn: sqlite3.Connection):
    """Extend schema with accounting tables. Safe to call multiple times."""
    for stmt in ACCOUNTING_SCHEMA.strip().split(";"):
        stmt = stmt.strip()
        if not stmt:
            continue
        try:
            conn.execute(stmt)
        except sqlite3.OperationalError as e:
            if "duplicate column" in str(e).lower():
                pass  # ALTER TABLE already ran
            else:
                raise
    conn.commit()


# ── Invoice CRUD ──────────────────────────────────────────────────────────────
def create_invoice(order_id: str, customer_id: int, total_charged: float,
                   payment_method: str = None, due_days: int = 3) -> str:
    inv_id = str(uuid.uuid4())
    now    = datetime.datetime.utcnow()
    issued = now.strftime("%Y-%m-%d")
    due    = (now + datetime.timedelta(days=due_days)).strftime("%Y-%m-%d")
    with get_db() as conn:
        conn.execute(
            """INSERT INTO invoices
               (invoice_id, order_id, customer_id, issued_date, due_date,
                total_charged, amount_paid, payment_method, status)
               VALUES (?,?,?,?,?,?,0,?,'unpaid')""",
            (inv_id, order_id, customer_id, issued, due, total_charged, payment_method)
        )
    return inv_id


def mark_invoice_paid(invoice_id: str, amount_paid: float,
                      payment_method: str = None, payment_ref: str = None):
    paid_date = datetime.datetime.utcnow().strftime("%Y-%m-%d")
    with get_db() as conn:
        inv = conn.execute(
            "SELECT total_charged FROM invoices WHERE invoice_id = ?", (invoice_id,)
        ).fetchone()
        if not inv:
            raise ValueError(f"Invoice {invoice_id} not found")
        total = inv["total_charged"]
        status = "paid" if amount_paid >= total else "partial"
        conn.execute(
            """UPDATE invoices
               SET amount_paid=?, payment_method=COALESCE(?,payment_method),
                   paid_date=?, status=?
               WHERE invoice_id=?""",
            (amount_paid, payment_method, paid_date, status, invoice_id)
        )
        conn.execute(
            """UPDATE orders SET payment_ref=? WHERE id=(
               SELECT order_id FROM invoices WHERE invoice_id=?)""",
            (payment_ref, invoice_id)
        )


def get_invoice_by_order(order_id: str) -> Optional[dict]:
    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM invoices WHERE order_id = ? ORDER BY issued_date DESC LIMIT 1",
            (order_id,)
        ).fetchone()
    return dict(row) if row else None


# ── Income Statement ──────────────────────────────────────────────────────────
def compute_income_statement(period_start: str, period_end: str,
                              platform_fees: float = 0.0,
                              supply_expense: float = 0.0,
                              save: bool = True) -> dict:
    """
    Compute income statement for period_start..period_end (inclusive).
    Reads from: invoices (revenue), ledger (costs).
    """
    with get_db() as conn:
        # Revenue: paid invoices in period
        rev_row = conn.execute(
            """SELECT COALESCE(SUM(amount_paid), 0) as gross_revenue
               FROM invoices
               WHERE status IN ('paid','partial')
               AND paid_date BETWEEN ? AND ?""",
            (period_start, period_end)
        ).fetchone()
        gross_revenue = rev_row["gross_revenue"]

        # COGS: from ledger entries in period
        cogs_row = conn.execute(
            """SELECT
               COALESCE(SUM(material_cost),  0) as material_cogs,
               COALESCE(SUM(machine_cost),   0) as machine_cogs,
               COALESCE(SUM(labor_cost),     0) as labor_cogs,
               COALESCE(SUM(customization_cost), 0) as design_cogs
               FROM ledger
               WHERE date BETWEEN ? AND ?""",
            (period_start, period_end)
        ).fetchone()

    mat  = round(cogs_row["material_cogs"], 2)
    mach = round(cogs_row["machine_cogs"],  2)
    lab  = round(cogs_row["labor_cogs"],    2)
    des  = round(cogs_row["design_cogs"],   2)
    total_cogs = round(mat + mach + lab + des, 2)
    gross_profit = round(gross_revenue - total_cogs, 2)
    gross_margin = round((gross_profit / gross_revenue * 100) if gross_revenue else 0, 2)
    op_expense = round(platform_fees + supply_expense, 2)
    net_income = round(gross_profit - op_expense, 2)

    stmt = {
        "period_start":     period_start,
        "period_end":       period_end,
        "gross_revenue":    round(gross_revenue, 2),
        "material_cogs":    mat,
        "machine_cogs":     mach,
        "labor_cogs":       lab,
        "design_cogs":      des,
        "total_cogs":       total_cogs,
        "gross_profit":     gross_profit,
        "gross_margin_pct": gross_margin,
        "platform_fees":    round(platform_fees, 2),
        "supply_expense":   round(supply_expense, 2),
        "operating_expense": op_expense,
        "net_income":       net_income,
    }

    if save:
        now = datetime.datetime.utcnow().isoformat() + "Z"
        with get_db() as conn:
            conn.execute(
                """INSERT INTO income_statements
                   (period_start, period_end, gross_revenue,
                    material_cogs, machine_cogs, labor_cogs, design_cogs,
                    total_cogs, gross_profit, gross_margin_pct,
                    platform_fees, supply_expense, operating_expense,
                    net_income, generated_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (period_start, period_end, stmt["gross_revenue"],
                 mat, mach, lab, des, total_cogs, gross_profit, gross_margin,
                 platform_fees, supply_expense, op_expense, net_income, now)
            )
    return stmt


# ── Balance Sheet ─────────────────────────────────────────────────────────────
def _equipment_value(snapshot_date: str) -> float:
    purchase = datetime.date.fromisoformat(PRINTER_PURCHASE_DATE)
    snap     = datetime.date.fromisoformat(snapshot_date)
    years    = (snap - purchase).days / 365.25
    annual_dep = PRINTER_PURCHASE_PRICE / PRINTER_USEFUL_LIFE_YR
    value = max(0.0, PRINTER_PURCHASE_PRICE - annual_dep * years)
    return round(value, 2)


def compute_balance_sheet(snapshot_date: str = None,
                           cash: float = 0.0,
                           inventory_grams: float = 0.0,
                           filament_cost_per_gram: float = 0.022,
                           accounts_payable: float = 0.0,
                           save: bool = True) -> dict:
    """
    Compute balance sheet snapshot.
    cash, inventory_grams, and accounts_payable must be supplied by owner
    (system can't know what's in your Zelle/CashApp or how much filament is on the shelf).
    """
    if snapshot_date is None:
        snapshot_date = datetime.date.today().isoformat()

    with get_db() as conn:
        ar_row = conn.execute(
            """SELECT COALESCE(SUM(total_charged - amount_paid), 0) as ar
               FROM invoices
               WHERE status NOT IN ('paid','void','refunded')"""
        ).fetchone()
    accounts_receivable = round(ar_row["ar"], 2)
    inventory_value     = round(inventory_grams * filament_cost_per_gram, 2)
    equip_value         = _equipment_value(snapshot_date)
    total_assets        = round(cash + accounts_receivable + inventory_value + equip_value, 2)
    total_liabilities   = round(accounts_payable, 2)
    owners_equity       = round(total_assets - total_liabilities, 2)

    bs = {
        "snapshot_date":       snapshot_date,
        "cash":                round(cash, 2),
        "accounts_receivable": accounts_receivable,
        "inventory_value":     inventory_value,
        "equipment_value":     equip_value,
        "total_assets":        total_assets,
        "accounts_payable":    round(accounts_payable, 2),
        "total_liabilities":   total_liabilities,
        "owners_equity":       owners_equity,
    }

    if save:
        now = datetime.datetime.utcnow().isoformat() + "Z"
        with get_db() as conn:
            conn.execute(
                """INSERT INTO balance_sheets
                   (snapshot_date, cash, accounts_receivable, inventory_value,
                    equipment_value, total_assets, accounts_payable,
                    total_liabilities, owners_equity, generated_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (snapshot_date, bs["cash"], accounts_receivable, inventory_value,
                 equip_value, total_assets, accounts_payable,
                 total_liabilities, owners_equity, now)
            )
    return bs


# ── Monthly Report ────────────────────────────────────────────────────────────
def compute_monthly_report(year: int, month: int,
                            platform_fees: float = 0.0,
                            supply_expense: float = 0.0,
                            save: bool = True) -> dict:
    month_str   = f"{year:04d}-{month:02d}"
    period_start = f"{month_str}-01"
    # Last day of month
    if month == 12:
        period_end = f"{year}-12-31"
    else:
        period_end = (datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)).isoformat()

    with get_db() as conn:
        # Jobs completed / failed in month
        jobs_row = conn.execute(
            """SELECT
               SUM(CASE WHEN status='complete' THEN 1 ELSE 0 END) as completed,
               SUM(CASE WHEN status='failed'   THEN 1 ELSE 0 END) as failed
               FROM orders
               WHERE completed_at BETWEEN ? AND ?""",
            (period_start + "T00:00:00Z", period_end + "T23:59:59Z")
        ).fetchone()
        jobs_completed = jobs_row["completed"] or 0
        jobs_failed    = jobs_row["failed"] or 0
        total_jobs     = jobs_completed + jobs_failed
        failure_rate   = round((jobs_failed / total_jobs * 100) if total_jobs else 0, 2)

        # Revenue (paid invoices)
        rev_row = conn.execute(
            """SELECT COALESCE(SUM(amount_paid), 0) as revenue
               FROM invoices
               WHERE status IN ('paid','partial')
               AND paid_date BETWEEN ? AND ?""",
            (period_start, period_end)
        ).fetchone()
        gross_revenue = round(rev_row["revenue"], 2)

        # COGS
        cogs_row = conn.execute(
            """SELECT
               COALESCE(SUM(material_cost + machine_cost + labor_cost + customization_cost), 0) as total_cogs,
               COALESCE(SUM(0), 0) as filament
               FROM ledger WHERE date BETWEEN ? AND ?""",
            (period_start, period_end)
        ).fetchone()
        total_cogs = round(cogs_row["total_cogs"], 2)

        # Machine hours (from order items, if tracked — fallback: ledger machine_cost / rate)
        hours_row = conn.execute(
            """SELECT COALESCE(SUM(machine_cost), 0) as mach_cost
               FROM ledger WHERE date BETWEEN ? AND ?""",
            (period_start, period_end)
        ).fetchone()
        machine_hours = round(hours_row["mach_cost"] / MACHINE_RATE_PER_HOUR, 2)

        # Filament used: machine_cost / 0.20 gives hours, but we want grams
        # Pull from order_items if available
        fil_row = conn.execute(
            """SELECT COALESCE(SUM(oi.quantity *
               CAST(json_extract(oi.customizations, '$.filament_grams') AS FLOAT)), 0) as grams
               FROM order_items oi
               JOIN orders o ON o.id = oi.order_id
               WHERE o.completed_at BETWEEN ? AND ?""",
            (period_start + "T00:00:00Z", period_end + "T23:59:59Z")
        ).fetchone()
        filament_used = round(fil_row["grams"] or 0, 2)

        # New customers in month
        new_cust = conn.execute(
            """SELECT COUNT(*) as n FROM customers
               WHERE created_at BETWEEN ? AND ?""",
            (period_start + "T00:00:00Z", period_end + "T23:59:59Z")
        ).fetchone()["n"]

        # Repeat customers who ordered in month
        repeat_cust = conn.execute(
            """SELECT COUNT(DISTINCT o.customer_id) as n
               FROM orders o
               JOIN customers c ON c.id = o.customer_id
               WHERE o.created_at BETWEEN ? AND ?
               AND c.order_count > 1""",
            (period_start + "T00:00:00Z", period_end + "T23:59:59Z")
        ).fetchone()["n"]

        # Top customer
        top_row = conn.execute(
            """SELECT i.customer_id, SUM(i.amount_paid) as total
               FROM invoices i
               WHERE i.paid_date BETWEEN ? AND ?
               AND i.status IN ('paid','partial')
               GROUP BY i.customer_id
               ORDER BY total DESC LIMIT 1""",
            (period_start, period_end)
        ).fetchone()
        top_customer = top_row["customer_id"] if top_row else None

    gross_profit = round(gross_revenue - total_cogs, 2)
    gross_margin = round((gross_profit / gross_revenue * 100) if gross_revenue else 0, 2)
    op_expense   = round(platform_fees + supply_expense, 2)
    net_income   = round(gross_profit - op_expense, 2)
    avg_job_val  = round(gross_revenue / jobs_completed if jobs_completed else 0, 2)

    report = {
        "report_month":        month_str,
        "jobs_completed":      jobs_completed,
        "jobs_failed":         jobs_failed,
        "failure_rate_pct":    failure_rate,
        "gross_revenue":       gross_revenue,
        "total_cogs":          total_cogs,
        "gross_profit":        gross_profit,
        "gross_margin_pct":    gross_margin,
        "net_income":          net_income,
        "new_customers":       new_cust,
        "repeat_customers":    repeat_cust,
        "avg_job_value":       avg_job_val,
        "filament_used_grams": filament_used,
        "machine_hours":       machine_hours,
        "top_customer_id":     top_customer,
    }

    if save:
        now = datetime.datetime.utcnow().isoformat() + "Z"
        with get_db() as conn:
            conn.execute(
                """INSERT OR REPLACE INTO monthly_reports
                   (report_month, jobs_completed, jobs_failed, failure_rate_pct,
                    gross_revenue, total_cogs, gross_profit, gross_margin_pct,
                    net_income, new_customers, repeat_customers, avg_job_value,
                    filament_used_grams, machine_hours, top_customer_id, generated_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (month_str, jobs_completed, jobs_failed, failure_rate,
                 gross_revenue, total_cogs, gross_profit, gross_margin,
                 net_income, new_cust, repeat_cust, avg_job_val,
                 filament_used, machine_hours, top_customer, now)
            )
    return report


# ── Customer Ledger ───────────────────────────────────────────────────────────
def get_customer_ledger(customer_id: int = None) -> list:
    """
    Returns per-customer accounting view: orders, amounts, outstanding balance.
    Decrypts PII fields.
    """
    with get_db() as conn:
        if customer_id:
            rows = conn.execute(
                "SELECT * FROM customers WHERE id = ?", (customer_id,)
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM customers ORDER BY id").fetchall()

        ledger = []
        for r in rows:
            cid = r["id"]
            invoices = conn.execute(
                """SELECT o.id as order_id, i.invoice_id, i.issued_date,
                          i.total_charged, i.amount_paid, i.status, i.payment_method
                   FROM invoices i JOIN orders o ON o.id = i.order_id
                   WHERE i.customer_id = ? ORDER BY i.issued_date""",
                (cid,)
            ).fetchall()

            outstanding = sum(
                (i["total_charged"] - i["amount_paid"])
                for i in invoices
                if i["status"] not in ("paid", "void", "refunded")
            )
            lifetime = sum(i["amount_paid"] for i in invoices)

            rd = dict(r)
            ledger.append({
                "customer_id":       cid,
                "name":              decrypt(rd["name"]),
                "discord_id":        rd["discord_id"],
                "channel":           rd.get("channel", "discord"),
                "status":            rd.get("customer_status", "active"),
                "total_jobs":        rd["order_count"],
                "lifetime_revenue":  round(lifetime, 2),
                "outstanding_balance": round(outstanding, 2),
                "invoices":          [dict(i) for i in invoices],
            })
    return ledger


# ── Exports ───────────────────────────────────────────────────────────────────
def export_csv(period: str) -> dict:
    """
    Export all accounting data for a period (YYYY-MM or YYYY) to CSV files.
    Returns dict of {table_name: filepath}.
    period = '2026-05' → one month. period = '2026' → full year.
    """
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    files = {}

    with get_db() as conn:
        for table, query, fname in [
            ("invoices",
             f"SELECT * FROM invoices WHERE substr(issued_date,1,{len(period)})='{period}'",
             f"{period}_invoices.csv"),
            ("ledger",
             f"SELECT * FROM ledger WHERE substr(date,1,{len(period)})='{period}'",
             f"{period}_ledger.csv"),
            ("monthly_reports",
             f"SELECT * FROM monthly_reports WHERE substr(report_month,1,{len(period)})='{period}'",
             f"{period}_monthly_reports.csv"),
        ]:
            rows = conn.execute(query).fetchall()
            if not rows:
                continue
            path = EXPORTS_DIR / fname
            with open(path, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=rows[0].keys())
                writer.writeheader()
                writer.writerows([dict(r) for r in rows])
            files[table] = str(path)

    return files


def export_xlsx(period: str,
                platform_fees: float = 0.0,
                supply_expense: float = 0.0) -> str:
    """
    Generate a formatted .xlsx workbook for the period.
    Sheets: Summary, Income Statement, Ledger, Customers, Invoices.
    Returns filepath.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
    except ImportError:
        raise RuntimeError("openpyxl not installed — run: pip install openpyxl")

    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # Determine period bounds
    if len(period) == 7:   # YYYY-MM
        y, m = int(period[:4]), int(period[5:7])
        p_start = f"{period}-01"
        if m == 12:
            p_end = f"{y}-12-31"
        else:
            p_end = (datetime.date(y, m + 1, 1) - datetime.timedelta(days=1)).isoformat()
    else:                  # YYYY
        p_start = f"{period}-01-01"
        p_end   = f"{period}-12-31"

    _HEADER = Font(bold=True, color="FFFFFF")
    _BG     = PatternFill("solid", fgColor="1F3864")
    _BG2    = PatternFill("solid", fgColor="2E4D8B")
    _MONEY  = '#,##0.00'

    def _header_row(ws, cols):
        ws.append(cols)
        for cell in ws[ws.max_row]:
            cell.font = _HEADER
            cell.fill = _BG
            cell.alignment = Alignment(horizontal="center")

    def _money_col(ws, col_letter, start_row, end_row):
        for row in range(start_row, end_row + 1):
            ws[f"{col_letter}{row}"].number_format = _MONEY

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    stmt = compute_income_statement(p_start, p_end, platform_fees, supply_expense, save=False)

    ws.append([f"Typhon's Forge — {period}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    rows_s = [
        ("Gross Revenue",         stmt["gross_revenue"]),
        ("Material COGS",         stmt["material_cogs"]),
        ("Machine COGS",          stmt["machine_cogs"]),
        ("Labor COGS",            stmt["labor_cogs"]),
        ("Design COGS",           stmt["design_cogs"]),
        ("Total COGS",            stmt["total_cogs"]),
        ("Gross Profit",          stmt["gross_profit"]),
        ("Gross Margin %",        stmt["gross_margin_pct"]),
        ("Platform Fees",         stmt["platform_fees"]),
        ("Supply Expense",        stmt["supply_expense"]),
        ("Net Income",            stmt["net_income"]),
    ]
    for label, val in rows_s:
        ws.append([label, val])
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = _MONEY
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14

    # ── Sheet 2: Income Statement ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Income Statement")
    _header_row(ws2, list(stmt.keys()))
    ws2.append(list(stmt.values()))

    # ── Sheet 3: Ledger ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Ledger")
    with get_db() as conn:
        ledger_rows = conn.execute(
            f"SELECT * FROM ledger WHERE date BETWEEN ? AND ? ORDER BY date",
            (p_start, p_end)
        ).fetchall()
    if ledger_rows:
        _header_row(ws3, list(ledger_rows[0].keys()))
        for r in ledger_rows:
            ws3.append(list(r))

    # ── Sheet 4: Invoices ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Invoices")
    with get_db() as conn:
        inv_rows = conn.execute(
            f"SELECT * FROM invoices WHERE issued_date BETWEEN ? AND ? ORDER BY issued_date",
            (p_start, p_end)
        ).fetchall()
    if inv_rows:
        _header_row(ws4, list(inv_rows[0].keys()))
        for r in inv_rows:
            ws4.append(list(r))

    # ── Sheet 5: Customers ────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Customers")
    ledger_data = get_customer_ledger()
    if ledger_data:
        _header_row(ws5, ["customer_id","name","discord","channel",
                           "status","total_jobs","lifetime_revenue","outstanding_balance"])
        for c in ledger_data:
            ws5.append([
                c["customer_id"], c["name"], c["discord_id"],
                c["channel"], c["status"], c["total_jobs"],
                c["lifetime_revenue"], c["outstanding_balance"],
            ])

    fname = f"{period}_TyphonsForge.xlsx"
    path  = EXPORTS_DIR / fname
    wb.save(str(path))
    return str(path)


# ── Plain-text dashboard summary ──────────────────────────────────────────────
def dashboard_summary(year: int = None, month: int = None) -> str:
    """Quick text summary for the Flask dashboard header card."""
    now  = datetime.date.today()
    y    = year  or now.year
    m    = month or now.month
    p    = f"{y:04d}-{m:02d}"
    p_s  = f"{p}-01"
    p_e  = (datetime.date(y, m + 1, 1) - datetime.timedelta(days=1)).isoformat() \
           if m < 12 else f"{y}-12-31"

    stmt = compute_income_statement(p_s, p_e, save=False)

    with get_db() as conn:
        order_count = conn.execute(
            "SELECT COUNT(*) FROM orders WHERE status='complete' AND completed_at BETWEEN ? AND ?",
            (p_s + "T00:00:00Z", p_e + "T23:59:59Z")
        ).fetchone()[0]
        pending_count = conn.execute(
            "SELECT COUNT(*) FROM orders WHERE status IN ('quoted','pending_payment','paid')"
        ).fetchone()[0]

    lines = [
        f"{'─'*40}",
        f"  {p}  —  Typhon's Forge",
        f"{'─'*40}",
        f"  Revenue:    ${stmt['gross_revenue']:>8.2f}",
        f"  COGS:       ${stmt['total_cogs']:>8.2f}",
        f"  Gross:      ${stmt['gross_profit']:>8.2f}  ({stmt['gross_margin_pct']:.1f}%)",
        f"  Net income: ${stmt['net_income']:>8.2f}",
        f"{'─'*40}",
        f"  Completed orders:  {order_count}",
        f"  Pending queue:     {pending_count}",
        f"{'─'*40}",
    ]
    return "\n".join(lines)


if __name__ == "__main__":
    from shop.db import init_db, _connect
    init_db()
    conn = _connect()
    init_accounting(conn)
    conn.close()
    print(dashboard_summary())
    print("\nAccounting tables initialized.")
